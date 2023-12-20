import logging
import openpyxl as op
import time
from openpyxl.styles import Color, PatternFill, Font, Border
import os

# python version - Python 3.10.0

Testing_file_name = "Testing"
Loantape_file_name = "Loantape"

timestamp = str(time.strftime("%Y%m%d-%H%M%S"))
resultpath = r'C:\Users\ljurkowski001\PycharmProjects\VWcomparison\result_'+timestamp
os.makedirs(resultpath)
output_name = resultpath+"\output"+timestamp+".txt"
testing_name = resultpath + "\\" + Testing_file_name +timestamp+".xlsx"
loantape_name = resultpath+"\\"+Loantape_file_name+timestamp+".xlsx"
result_name = resultpath+"\Results"+timestamp+".xlsx"


# TO DO

# make sure that TESTING table has exactly the same column order as mapping tupple
# ID of the table has to be on the first place


logging.basicConfig(filename=output_name, filemode ='a', format='%(asctime)s,%(msecs)d %(name)s %(levelname)s %(message)s',datefmt='%H:%M:%S', level=logging.INFO)
white = PatternFill(fill_type =None)
not_found = PatternFill(fill_type ="solid", start_color="D79707", end_color="D79707")
green = PatternFill(fill_type ="solid", start_color="20BE60", end_color="20BE60")
red = PatternFill(fill_type = "solid", start_color="ee6b6e", end_color="ee6b6e")

def search_in_tuple(header_title):
    x = 0
    y = 0
    while (x < len(mapping_tuple)):
        if str(mapping_tuple[x][0]) == str(header_title):
            return str(mapping_tuple[x][1])
            break
        else:
            x = x + 1
        # print(mapping_tuple[x][y-1])
def find_column_number_in_loantape(column_name):
    list_index = 1
    logging.info("list = " + str(loantape_column_names) )
    while (list_index <= len(loantape_column_names)+1):
        if str(loantape_column_names[list_index-1]) == str(column_name):
            return list_index-1
            break
        else:
            list_index = list_index + 1
def looking_for_value_in_column(key_value, column_index):
    column_name = loantape_column_names[column_index]
    found_matches = 0
    return_tuple = (False, "", found_matches)
    for column_cell in loantape_sh.iter_cols(1, loantape_sh.max_row):
        if column_cell[0].value == column_name:
            for data in column_cell[1:]:
                # print("XXXXXXXXXXXXXX  "+ data.value)
                if data.value is not None:
                    if data.value == key_value:
                        logging.debug("I've found it in column " + column_name + " " + str(data.coordinate))
                        data.fill = green
                        found_matches = found_matches + 1
                        logging.info(return_tuple)
                        return_tuple = (True, str(data.row), found_matches)
                else:
                    logging.info("I've reached eof")
                    break
    return return_tuple


    return return_tuple
def read_value_in_column_and_row(row, column):
    logging.debug("I'm in read_value_in_column_and_row and looking for: " + str(loantape_sh.cell(row=int(row), column=int(column)).value))
    return str(loantape_sh.cell(row=row, column=column).value)

# stores number of rows where all instr_id taken from testing file was not found in second table
not_found_key = 0
# stores number of rows where all of values matches between files
all_correct_values = 0
lessorequal5fails = 0
over5fails = 0
over10fails = 0
over20fails = 0
duplicated_key = 0
# number_of_rows_testing = 0
# number_of_rows_loantape = 0


# 1. read file testing
testing_worksheet = op.load_workbook(Testing_file_name+".xlsx")
testing_sh = testing_worksheet.worksheets[0]
# 2. read file loantape
loantape_worksheet = op.load_workbook(Loantape_file_name+".xlsx")
loantape_sh = loantape_worksheet.worksheets[0]

# 3. define mapping tuple - order should be the same as in testing table
#
mapping_tuple = (("KDW0001", "BIC/KLL1"), ("KDW0002", "BIC/LK2"), ("KDW0003", "BIC/77382"),("KDW0004", "/BIC/HKSDI2"), ("KDW0005", "/BIC/70836"), ("KDW0006", "/BIC/SHUDIUS8"), ("KDW0007", "/BIC/HYUJ80"), ("KDW0008", "/BIC/POIUY"), ("KDW0009", "/BIC/ASDF"), ("KDW00010", "/BIC/00908"))
# 4. check number of rows
row_number_testing = testing_sh.max_row
row_number_loantape = loantape_sh.max_row
logging.info("testing rows no: " + str(row_number_testing) + " ; loantape rows no: " + str(row_number_loantape))

# go through loantape column list and save it
loantape_column_names = list()
for column_cell in loantape_sh.iter_cols():
    if str(column_cell[0].value) is not None:
        loantape_column_names.append(str(column_cell[0].value))


# print("max col " + str(loantape_sh.max_column))
# for column_value in loantape_sh.iter_cols(1,):
#      if column_value[0].value is None:
#         for cell in column_value[1:]:
#             cell.value = ""

column_numbers = len(loantape_column_names)
# print("len(loantape_column_names) " + str(column_numbers) + " mapping tuple len = " + str(len(mapping_tuple)))
# print("max_column " + str(loantape_sh.max_column))
loantape_sh.insert_cols(idx=int(str(loantape_sh.max_column+1)))
# print("max_column after added " + str(loantape_sh.max_column))
max_column = loantape_sh.max_column
loantape_sh.delete_cols(idx=int(column_numbers+1))

key_value = "A1"
i = 2

# while (i <= row_number_testing):
for row_cell in testing_sh.iter_rows(2,):
    if row_cell[0].value is not None:
        logging.info("***************************************** ROW " + str(i) + "*****************************************")
        key_column_name = testing_sh[key_value].value
        column_returned = search_in_tuple(testing_sh[key_value].value)

        # I'm looking now in loantape table column with that name
        index_loantape_column_names = find_column_number_in_loantape(column_returned)
        our_key_index = "A"+str(i)
        our_key_value = testing_sh[our_key_index].value

        logging.info("Current value I'll be looking at loantape table as key:  " + str(our_key_value))
        key_value_tuple = (False, "A1")
        key_value_tuple = looking_for_value_in_column(our_key_value, index_loantape_column_names)
        # print("key_value_tuple: " + str(key_value_tuple[0]) + str(key_value_tuple[1]) + str(key_value_tuple[2]))


        if key_value_tuple is not None and key_value_tuple[0] == True and key_value_tuple[2] == 1:
            logging.debug("I've found the key value in loantape column")
            # I've found the key, now I can go through the rest of columns
            local_match = 0
            local_fail = 0
            local_errors = ""
            column_index = 2
            for cell in row_cell[1:max_column]:
                # if cell.column <= max_column and cell.value is not None:
                logging.info("----------------------------")

                logging.debug("matched row_testing: " + str(i) + " and row_loantape: " + key_value_tuple[1])
                # make sure that TESTING table has exactly the same column order as mapping tupple

                # find value in following column and in above row                    logging.debug(str(mapping_tuple[column_index-1][1]))
                # 1. read the column_header of that value from tuple
                # print(column_index)
                index_column = find_column_number_in_loantape(mapping_tuple[column_index-1][1])
                # print ("read value= " + str(loantape_sh.cell(row=5, column=10).value))
                returned_value = loantape_sh.cell(row=int(key_value_tuple[1]), column=int(index_column+1)).value
                logging.info("testing value = " + str(cell.value) + "; loantape value = " + str(returned_value))
                # check empty value scenario
                loantape_sh.cell(row=int(key_value_tuple[1]), column=int(index_column+1)).fill = white


                if cell.value is None:
                    cell.value = ""
                if returned_value is None:
                    returned_value = ""
                if cell.value == returned_value:
                    local_match=local_match+1
                    loantape_sh.cell(row=int(key_value_tuple[1]), column=int(index_column+1)).fill = green
                    cell.fill = green
                    logging.info("We've got a match. Local_match = " + str(local_match) + " and local_fail = " + str(local_fail))
                else:
                    local_fail=local_fail+1

                    loantape_sh.cell(row=int(key_value_tuple[1]), column=int(index_column+1)).fill = red
                    cell.fill = red

                    loantape_sh.cell(row=int(key_value_tuple[1]), column=int(index_column+1)).font = Font(bold=True)
                    logging.info("Values not matched. Local_match = " + str(local_match) + " and local_fail = " + str(local_fail))
                    local_errors = local_errors + "In column=" + str(mapping_tuple[column_index-1][1]) + ", Testing value=" + cell.value + ", Loantape value=" + returned_value + "; "

                column_index = column_index + 1

            if local_fail > 0:
                # print(local_errors)
                loantape_sh.cell(row=int(key_value_tuple[1]),column=max_column+1).value = "Number of failures: " + str(local_fail) + "; " + str(local_errors)


        elif key_value_tuple is not None and key_value_tuple[2] == 0:
            logging.debug("I've not found it")
            testing_sh.cell(row=i, column=max_column+1).value = "Key not found"
            row_cell[0].fill = not_found
            not_found_key = not_found_key + 1
            local_match = 0

        elif key_value_tuple is not None and key_value_tuple[2] > 1:
            logging.debug("That key has more than one match")
            duplicated_key = duplicated_key + 1
            testing_sh.cell(row=i, column=int(max_column+1)).value = "Testing value: " + str(our_key_value) + " has more than one match"





        logging.info("for entire row " + str(i) + " , matched values: " + str(local_match) + " and not matched values: " + str(local_fail))
        if local_match > 0 and local_fail == 0 :
            all_correct_values = all_correct_values + 1
        elif local_fail > 0 and local_fail <= 5:
            lessorequal5fails = lessorequal5fails + 1
        elif local_fail > 5 and local_fail <= 10:
            over5fails = over5fails + 1
        elif local_fail > 10 and local_fail <= 20:
            over10fails = over10fails + 1
        elif local_fail > 20:
            over20fails = over20fails + 1

        logging.debug("not found key rows: " + str(not_found_key))

        i= i+1

    else:
        logging.info("I've reached end of file in row " + str(i))
        break

loantape_worksheet.save(loantape_name)
testing_worksheet.save(testing_name)

# to add also the percentage of the results based on the numbers

logging.info("Not found keys: " + str(not_found_key) + "; " + str(round(100* not_found_key/row_number_testing, 2)) + "%")
logging.info("Duplicated keys: " + str(duplicated_key) + "; " + str(round(100* duplicated_key/row_number_testing, 2)) + "%")
logging.info("rows with all correct values: " + str(all_correct_values) + "; " + str(round(100* all_correct_values/row_number_testing, 2)) + "%")
logging.info("rows with no more than 5 fails " + str(lessorequal5fails) + "; " + str(round(100* lessorequal5fails/row_number_testing, 2)) + "%")
logging.info("rows with no more than 10 fails " + str(over5fails) + "; " + str(round(100* over5fails/row_number_testing, 2)) + "%")
logging.info("rows with no more than 20 fails " + str(over10fails) + "; " + str(round(100* over10fails/row_number_testing, 2)) + "%")
logging.info("rows with more than 20 fails " + str(over20fails) + "; " + str(round(100* over20fails/row_number_testing, 2)) + "%")

res = op.Workbook()
res_sheet = res.active

for column_loantape in loantape_sh.iter_cols(1, column_numbers):
    # print(column_loantape[0].value)
    if str(column_loantape[0].value) is not None:
        number_of_errors = 0
        res_sheet.cell(row=1, column = int(column_loantape[0].column)).value = column_loantape[0].value
        for data in column_loantape[1:row_number_loantape]:
            # if data.fill.start_color.index == "ee6b6e":
            # print("data = " + str(data.value))
            # print("colour = " + str(data.fill.start_color.index))
            if data.fill.start_color.index == "00ee6b6e":
                number_of_errors = number_of_errors + 1
                # print("I'm in column " + str(column_loantape[0].value) +  " and row " + str(data.row) + "and value checked = " + str(data.value) + " and colour = " + str(data.fill.start_color.index))
        res_sheet.cell(row=2, column=int(column_loantape[0].column)).value =number_of_errors

res.save(result_name)
#